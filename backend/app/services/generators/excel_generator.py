import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def generate_excel(report_data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Заголовок
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = report_data.get("title", "Отчёт")
    title_cell.font = Font(bold=True, size=14, color="1677FF")
    title_cell.alignment = Alignment(horizontal="center")

    # Мета
    ws["A2"] = f"Период: {report_data.get('period_from', '—')} — {report_data.get('period_to', '—')}"
    ws["A3"] = f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A3"].font = Font(color="999999", italic=True)

    columns = report_data.get("columns", [])
    data = report_data.get("data", [])

    if columns:
        header_row = 5
        header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(col_name) + 4)

        for row_idx, row in enumerate(data, header_row + 1):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
